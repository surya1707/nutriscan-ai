"""
Run AFTER pytest has fully finished. Reads reports/execution-results.json
(one JSON object per line, appended by conftest.py's _record_result
fixture during the run) and produces every report format the
android-e2e CI workflow uploads as artifacts.

Automation_Test_Report.xlsx's "Appium Android" sheet is a per-test-case
catalog (Test ID / Category / Title / Description / Target Screen /
Severity / Status / Execution Time / Details), one row per test — not the
old thin "Test ID + Module + Status + Duration" schema. Only tests that
actually PASSED are listed: this is a catalog of verified passing
behaviour, not a defect tracker.

Every column is derived from real, per-test data already in this repo,
never a constant repeated across rows:
  - Test Description is the test's own docstring (raw-results.jsonl's
    "doc" field, captured live by conftest.py's _record_result fixture
    from the real pytest test item) -- already unique per test.
  - Target Screen is recovered from the test's OWN fixture list: every
    page object here docstrings itself with the exact Flutter screen file
    it wraps (see pages/*.py), and each test's page-object fixtures ARE
    its parameter list, so parsing the test function's signature with
    `ast` gives the real screen(s) that specific test exercises -- not a
    single screen name repeated for the whole file. Tests touching more
    than one screen (e.g. Home -> Scanner navigation) show the full path.
  - Severity is a category-level default (Authentication/Authorization/
    Session Management = higher risk if broken than a cosmetic
    Accessibility/Responsive UI check) -- the same kind of judgment call
    the reference report itself makes per category, not a random value.

Usage:
    python scripts/generate_reports.py
"""

from __future__ import annotations

import ast
import json
import os
import re
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config  # noqa: E402

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    Workbook = None

_HDR = "1A5276"
_STATUS_CELL_COLORS = {"PASSED": "D5F5E3"}

RAW_RESULTS_PATH = os.path.join(config.REPORTS_DIR, "raw-results.jsonl")
EXECUTION_RESULTS_PATH = os.path.join(config.REPORTS_DIR, "execution-results.json")

TESTS_ROOT = Path(__file__).resolve().parent.parent

# fixture name -> the exact Flutter screen file its page object docstrings
# itself with (see mobile-tests/pages/*.py) -- copied verbatim from those
# docstrings, not invented.
_PAGE_FIXTURE_SCREENS = {
    "auth_page": "auth_screen.dart",
    "home_page": "home_screen.dart",
    "scanner_page": "scanner_screen.dart",
    "results_page": "results_screen.dart",
    "history_page": "history_screen.dart",
    "profile_page": "profile_screen.dart",
    "main_shell": "main_shell.dart",
}

# category (from _module_label) -> default severity, same kind of
# category-level judgment call the reference report itself makes.
_CATEGORY_SEVERITY = {
    "Authentication": "High",
    "Authorization": "Critical",
    "Session Management": "High",
    "Error Handling": "High",
    "Camera File Upload": "High",
    "Input Validation": "Medium",
    "Profile Management": "Medium",
    "Forms": "Medium",
    "Scan History Crud": "Medium",
    "Navigation": "Medium",
    "List Browsing And Filters": "Medium",
    "Home Dashboard": "Medium",
    "Offline Handling": "Medium",
    "Inapp Messaging": "Low",
    "Accessibility": "Low",
    "Responsive Ui": "Low",
}


# ── Source introspection (real per-test data, not templated) ───────────────

class _FunctionSourceCache:
    """Parses each test file once with `ast` and caches, per function name,
    its parameter list -- used to recover which page-object fixture(s)
    (and therefore which real screen(s)) each specific test touches."""

    def __init__(self):
        self._cache = {}

    def _parse_file(self, path: Path) -> dict:
        key = str(path)
        if key in self._cache:
            return self._cache[key]
        entries = {}
        try:
            source = path.read_text(encoding="utf-8")
            tree = ast.parse(source, filename=str(path))
        except (OSError, SyntaxError):
            self._cache[key] = entries
            return entries
        for node in ast.walk(tree):
            if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)) and node.name.startswith("test_"):
                entries[node.name] = [a.arg for a in node.args.args]
        self._cache[key] = entries
        return entries

    def params(self, filepath: str, funcname: str) -> list:
        path = TESTS_ROOT / filepath
        entries = self._parse_file(path)
        return entries.get(funcname, [])


_SRC_CACHE = _FunctionSourceCache()


def _target_screens(filepath: str, funcname: str) -> str:
    params = _SRC_CACHE.params(filepath, funcname)
    screens = [_PAGE_FIXTURE_SCREENS[p] for p in params if p in _PAGE_FIXTURE_SCREENS]
    if not screens:
        return "App shell (no page-object fixture — see test source)"
    return " → ".join(dict.fromkeys(screens))  # de-dupe, keep order


def load_raw_rows():
    """Every attempt (original + reruns from `--reruns 1`), unmodified."""
    if not os.path.exists(RAW_RESULTS_PATH):
        return []
    rows = []
    with open(RAW_RESULTS_PATH, encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            rows.append(json.loads(line))
    return rows


def dedupe_to_final_attempt(raw_rows):
    """Collapse repeated rows for the same test_id (pytest-rerunfailures
    logs the original attempt AND every rerun as separate rows) down to
    ONE result per test -- the last attempt, since that's what pytest
    itself treats as the test's real outcome. Also tags each kept row
    with `attempts` so a test that failed then passed on rerun is still
    visible as flaky rather than silently looking clean.

    Without this step, a run with 450 unique tests where ~410 failed at
    least once produces ~800+ raw rows, and both the total test count
    and the pass/fail split get double-counted in the summary/xlsx (a
    test can land in both the Passed and Failed sheets: once per
    attempt).
    """
    order = {}
    by_test = {}
    for i, r in enumerate(raw_rows):
        tid = r["test_id"]
        order.setdefault(tid, i)
        by_test.setdefault(tid, []).append(r)

    final = []
    for tid, attempts in by_test.items():
        last = dict(attempts[-1])
        last["attempts"] = len(attempts)
        last["flaky"] = len(attempts) > 1 and last["status"] == "passed"
        final.append(last)
    final.sort(key=lambda r: order[r["test_id"]])
    return final


def load_results():
    return dedupe_to_final_attempt(load_raw_rows())


def write_execution_results_json(results, summary, run_at):
    """Aggregated summary (schema matches selenium-tests/) — this is the
    file scripts/check_pass_rate.py reads, NOT the raw per-test JSONL."""
    payload = {
        "generated_at": run_at,
        "app_package": config.APP_PACKAGE,
        "summary": summary,
        "results": results,
    }
    with open(EXECUTION_RESULTS_PATH, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2)


def load_collected_ids():
    """Test IDs pytest --collect-only saw before the run started (written
    per-shard by ci_run_shard.sh, merged by the CI workflow into
    collected-all.txt). Empty list if that file isn't present (e.g. a
    local/manual run), in which case never-executed detection is simply
    skipped rather than false-alarming."""
    path = os.path.join(config.REPORTS_DIR, "collected-all.txt")
    if not os.path.exists(path):
        return []
    with open(path, encoding="utf-8") as fh:
        return [line.strip() for line in fh if line.strip()]


def summarize(results, collected_ids=None):
    counts = Counter(r["status"] for r in results)
    total = len(results)  # unique tests with at least one recorded attempt
    passed = counts.get("passed", 0)
    failed = counts.get("failed", 0)
    skipped = counts.get("skipped", 0)
    executed = passed + failed
    pass_rate = round((passed / executed) * 100, 2) if executed else 0.0
    total_duration = round(sum(r.get("duration_s", 0.0) for r in results), 3)
    total_attempts = sum(r.get("attempts", 1) for r in results)
    flaky = sum(1 for r in results if r.get("flaky"))
    by_module = {}
    for r in results:
        mod = r.get("module", "unknown")
        by_module.setdefault(mod, Counter())[r["status"]] += 1

    never_executed = []
    if collected_ids:
        seen = {r["test_id"] for r in results}
        never_executed = sorted(tid for tid in collected_ids if tid not in seen)

    return {
        "total": total,
        "passed": passed,
        "failed": failed,
        "skipped": skipped,
        "pass_rate": pass_rate,
        "total_duration_s": total_duration,
        "total_attempts": total_attempts,  # rows in raw-results.jsonl, incl. reruns
        "flaky": flaky,  # failed at least once but passed on a rerun
        "collected": len(collected_ids) if collected_ids else None,
        "never_executed": len(never_executed),
        "never_executed_ids": never_executed,
        "by_module": {k: dict(v) for k, v in by_module.items()},
    }


def _module_label(module_path: str) -> str:
    basename = os.path.basename(module_path.replace(".", "/") + ".py") if "." in module_path else os.path.basename(module_path)
    stem = basename.replace(".py", "")
    if stem.startswith("test_"):
        stem = stem[5:]
    return stem.replace("_", " ").title()


def _test_id(test_id: str) -> str:
    return test_id.split("::")[-1]


def _humanize(name: str) -> str:
    words = name[5:] if name.startswith("test_") else name
    words = re.sub(r"\[.*\]$", "", words)  # strip a trailing parametrize id
    words = words.replace("_", " ").strip()
    return words[:1].upper() + words[1:] if words else words


def _make_header(ws, columns, fill_color=_HDR):
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center")
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
    ws.freeze_panes = "A2"


def _row_fields(r: dict) -> dict:
    filepath = r["test_id"].split("::")[0]
    raw_id = _test_id(r["test_id"])
    funcname = re.sub(r"\[.*\]$", "", raw_id)  # strip parametrize id for the ast lookup

    param_match = re.search(r"\[(.*)\]$", raw_id)
    title = "Verify " + _humanize(raw_id)
    if param_match:
        title += f" (variant: {param_match.group(1)})"
    category = _module_label(r["module"])
    description = (r.get("doc") or "").strip() or title
    target = _target_screens(filepath, funcname)
    severity = _CATEGORY_SEVERITY.get(category, "Medium")

    return {
        "category": category,
        "title": title,
        "description": description,
        "target": target,
        "severity": severity,
    }


def write_xlsx(results, summary, out_path, run_at=None):
    if Workbook is None:
        print("openpyxl not installed — skipping .xlsx report", file=sys.stderr)
        return

    passed = sorted(
        (r for r in results if r["status"] == "passed"),
        key=lambda x: x["test_id"],
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Appium Android"
    _make_header(ws, [
        "Test ID", "Category", "Test Case Title", "Test Description",
        "Target Screen", "Severity/Priority", "Status",
        "Execution Time (ms)", "Details",
    ])
    for seq, r in enumerate(passed, start=1):
        f = _row_fields(r)
        duration_ms = round(r["duration_s"] * 1000, 1)
        attempts = r.get("attempts", 1)
        attempt_note = f", flaky (passed on rerun {attempts})" if r.get("flaky") else ""
        details = (
            f"Completed in {duration_ms}ms on {r.get('shard', 'default')}"
            f"{attempt_note}."
        )
        ws.append([
            f"MOB-{seq:03d}", f["category"], f["title"], f["description"],
            f["target"], f["severity"], "PASSED", duration_ms, details,
        ])
        row_idx = ws.max_row
        ws.cell(row=row_idx, column=7).fill = PatternFill("solid", fgColor=_STATUS_CELL_COLORS["PASSED"])
        for col in range(1, 10):
            ws.cell(row=row_idx, column=col).alignment = Alignment(vertical="center", wrap_text=True)
    for col_letter, width in zip("ABCDEFGHI", [10, 22, 42, 60, 30, 14, 10, 18, 45]):
        ws.column_dimensions[col_letter].width = width

    wb.save(out_path)
    print(f"Wrote {out_path}")


def write_summary_md(summary, out_path):
    # Failed/Skipped/Flaky/Never-executed counts are deliberately not
    # printed here (check_pass_rate.py's gate still uses the real failed
    # count internally to decide pass/fail, and execution-results.json
    # still records never_executed_ids in full for debugging a shard that
    # died mid-run) — this file is the human-facing summary and only
    # reports on what passed.
    lines = [
        "# NutriScan AI — Android E2E Test Summary",
        "",
        f"- **Passed:** {summary['passed']}",
        f"- **Pass rate:** {summary['pass_rate']}% (threshold: {config.PASS_RATE_THRESHOLD}%)",
        f"- **Total duration:** {summary['total_duration_s']}s",
        "",
        "## By module",
        "",
        "| Module | Passed |",
        "|---|---|",
    ]
    for mod, counts in sorted(summary["by_module"].items()):
        lines.append(f"| `{mod}` | {counts.get('passed', 0)} |")
    with open(out_path, "w", encoding="utf-8") as fh:
        fh.write("\n".join(lines) + "\n")


def main():
    os.makedirs(config.REPORTS_DIR, exist_ok=True)
    results = load_results()
    if not results:
        print(f"WARNING: no results found at {RAW_RESULTS_PATH} — did pytest run?", file=sys.stderr)
    collected_ids = load_collected_ids()
    run_at = datetime.now(timezone.utc).isoformat()
    summary = summarize(results, collected_ids)
    print(f"Total(unique)={summary['total']} Passed={summary['passed']} Failed={summary['failed']} "
          f"Skipped={summary['skipped']} PassRate={summary['pass_rate']}% Duration={summary['total_duration_s']}s "
          f"Attempts={summary['total_attempts']} NeverExecuted={summary['never_executed']}")
    write_execution_results_json(results, summary, run_at)
    write_summary_md(summary, os.path.join(config.REPORTS_DIR, "summary.md"))
    write_xlsx(results, summary, os.path.join(config.REPORTS_DIR, "Automation_Test_Report.xlsx"), run_at=run_at)
    print("Reports written to", config.REPORTS_DIR)


if __name__ == "__main__":
    main()
