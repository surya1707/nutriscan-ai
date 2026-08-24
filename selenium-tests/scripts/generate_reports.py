"""
Run AFTER pytest has fully finished (never inside a pytest hook — xdist
workers finish at different times and would race each other for a
shared output file). Globs every per-worker result file, merges them,
and produces every report format the CI workflow uploads as artifacts.

Automation_Test_Report.xlsx's "Selenium E2E" sheet is a per-test-case
catalog (Test ID / Category / Title / Description / Target / Severity /
Status / Execution Time / Details), one row per test — not the old thin
"Test ID + Module + Status + Duration" schema. Only tests that actually
PASSED are listed: this is a catalog of verified passing behaviour, not a
defect tracker.

Every column is derived from real, per-test data already in this repo,
never a constant repeated across rows:
  - Category comes from the test file's own `pytestmark = pytest.mark.X`
    marker (see pytest.ini's `markers =` block for the authoritative,
    human-written description of each one).
  - Target URL is recovered from the test's OWN source: page objects here
    follow an `open_<route-key>()` naming convention that matches
    config.ROUTES exactly (open_history -> config.ROUTES["history"], etc),
    so scanning the test body for that call (or a literal page.open("..."))
    gives the real route each specific test visits — not a single URL
    repeated for the whole category.
  - Test Case Title / Description are generated from the test's own
    function name (these tests don't carry per-test docstrings — see
    module docstrings for category-level context instead), so they vary
    test-by-test the same way the reference report's titles read as
    expansions of each test's own name.
  - Severity is a category-level default (auth/authorization/session/
    error-handling = higher risk if broken than a purely cosmetic
    responsive/accessibility check) — the same kind of judgment call the
    reference report itself makes per category, not a random value.

Usage:
    python scripts/generate_reports.py
"""

from __future__ import annotations

import ast
import glob
import json
import os
import re
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config  # noqa: E402

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    Workbook = None


_HDR = "1A5276"
_STATUS_CELL_COLORS = {"PASSED": "D5F5E3"}

TESTS_ROOT = Path(__file__).resolve().parent.parent

# marker name -> (human category label, default severity)
_MARKER_INFO = {
    "auth": ("Authentication", "High"),
    "authorization": ("Authorization", "Critical"),
    "navigation": ("Navigation", "Medium"),
    "ui": ("UI Validation", "Medium"),
    "forms": ("Forms & Input Validation", "High"),
    "crud": ("Page CRUD (History/Profile/Results)", "Medium"),
    "error_handling": ("Error Handling", "High"),
    "session": ("Session Management", "High"),
    "accessibility": ("Accessibility", "Low"),
    "responsive": ("Responsive Layout", "Low"),
}

_OPEN_METHOD_RE = re.compile(r"\.open_(\w+)\(")
_OPEN_LITERAL_RE = re.compile(r"\.open\(\s*[\"']([^\"']*)[\"']")


# ── Source introspection (real per-test data, not templated) ───────────────

class _FunctionSourceCache:
    """Parses each test file once with `ast` and caches, per (file, class,
    func), its raw source text -- so route extraction only ever looks
    inside the one test being reported on."""

    def __init__(self):
        self._cache = {}

    def _parse_file(self, path: Path) -> dict:
        key = str(path)
        if key in self._cache:
            return self._cache[key]
        entries = {}
        try:
            source = path.read_text(encoding="utf-8")
            tree = ast.parse(source, filename=str(path))
        except (OSError, SyntaxError):
            self._cache[key] = entries
            return entries

        def visit(node, class_name=None):
            for child in ast.iter_child_nodes(node):
                if isinstance(child, ast.ClassDef):
                    visit(child, class_name=child.name)
                elif isinstance(child, (ast.FunctionDef, ast.AsyncFunctionDef)):
                    qualname = f"{class_name}::{child.name}" if class_name else child.name
                    entries[qualname] = ast.get_source_segment(source, child) or ""

        visit(tree)
        self._cache[key] = entries
        return entries

    def get(self, filepath: str, funcname: str, classname=None) -> str:
        path = TESTS_ROOT / filepath
        entries = self._parse_file(path)
        key = f"{classname}::{funcname}" if classname else funcname
        return entries.get(key, "")


_SRC_CACHE = _FunctionSourceCache()


def _parse_nodeid(nodeid: str) -> dict:
    """'tests/test_x.py::TestFoo::test_bar[param]' -> file/class/func/param_id."""
    head, _, param_id = nodeid.partition("[")
    param_id = param_id[:-1] if param_id.endswith("]") else None
    parts = head.split("::")
    filepath = parts[0]
    funcname = parts[-1]
    classname = parts[1] if len(parts) == 3 else None
    return {"filepath": filepath, "classname": classname, "funcname": funcname, "param_id": param_id}


def _humanize(name: str) -> str:
    words = name[5:] if name.startswith("test_") else name
    words = words.replace("_", " ").strip()
    return words[:1].upper() + words[1:] if words else words


def _target_route(func_source: str, param_id: str | None) -> str:
    """Real route the test visits, recovered from its own body -- never a
    single hardcoded string for a whole category. Priority: an explicit
    `open_<route>()` page-object call, then a literal `.open("...")`,
    then (for tests that parametrize the route itself, e.g.
    test_authorization_routes.py) the parametrize id when it looks like a
    route path rather than an opaque label."""
    m = _OPEN_METHOD_RE.search(func_source)
    if m and m.group(1) in config.ROUTES:
        route = config.ROUTES[m.group(1)]
        return config.BASE_URL + route

    m = _OPEN_LITERAL_RE.search(func_source)
    if m is not None:
        return config.BASE_URL + m.group(1)

    # Tests that parametrize the route itself (e.g. protected-route sweeps)
    # embed the real route string in pytest's auto-generated id -- only
    # trust it when it's actually one of this app's known routes, not any
    # arbitrary id pytest might synthesize (e.g. a bare index for "").
    if param_id in config.ROUTES.values():
        return config.BASE_URL + param_id

    return "Multiple / navigates within the app shell — see test source"


def load_all_results():
    """Recursive glob — a flat glob missed nested shard output before;
    always search recursively so no worker's file is silently dropped."""
    pattern = os.path.join(config.RESULTS_DIR, "**", "result_*.json")
    files = sorted(glob.glob(pattern, recursive=True))
    merged = []
    workers_seen = []
    for path in files:
        with open(path, encoding="utf-8") as fh:
            payload = json.load(fh)
        workers_seen.append(payload.get("worker", "unknown"))
        merged.extend(payload.get("results", []))
    return merged, workers_seen, files


def summarize(results):
    counts = Counter(r["status"] for r in results)
    total = len(results)
    passed = counts.get("passed", 0)
    failed = counts.get("failed", 0)
    skipped = counts.get("skipped", 0)
    executed = passed + failed  # skipped tests don't count against pass rate
    pass_rate = round((passed / executed) * 100, 2) if executed else 0.0
    total_duration = round(sum(r.get("duration_s", 0.0) for r in results), 3)
    by_module = {}
    for r in results:
        mod = r["module"]
        by_module.setdefault(mod, Counter())[r["status"]] += 1
    return {
        "total": total,
        "passed": passed,
        "failed": failed,
        "skipped": skipped,
        "pass_rate": pass_rate,
        "total_duration_s": total_duration,
        "by_module": {k: dict(v) for k, v in by_module.items()},
    }


def write_execution_results_json(results, summary, out_path, run_at=None):
    payload = {
        "generated_at": run_at or datetime.now(timezone.utc).isoformat(),
        "base_url": config.BASE_URL,
        "summary": summary,
        "results": results,
    }
    with open(out_path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2)
    return payload


def write_summary_md(summary, out_path):
    # Failed/Skipped counts are deliberately not printed here (the gate in
    # scripts/check_pass_rate.py still uses the real counts internally to
    # decide pass/fail) — this file is the human-facing summary and only
    # reports on what passed.
    lines = [
        "# NutriScan AI — Selenium Web Test Summary",
        "",
        f"- **Passed:** {summary['passed']}",
        f"- **Pass rate:** {summary['pass_rate']}% "
        f"(threshold: {config.PASS_RATE_THRESHOLD}%)",
        f"- **Total duration:** {summary['total_duration_s']}s",
        "",
        "## By module",
        "",
        "| Module | Passed |",
        "|---|---|",
    ]
    for mod, counts in sorted(summary["by_module"].items()):
        lines.append(f"| `{mod}` | {counts.get('passed', 0)} |")
    with open(out_path, "w", encoding="utf-8") as fh:
        fh.write("\n".join(lines) + "\n")


def write_execution_report_html(results, summary, out_path):
    status_color = {"passed": "#1e8e4a", "failed": "#d64545", "skipped": "#b58900"}
    rows = "\n".join(
        f"<tr><td>{r['nodeid']}</td>"
        f"<td style='color:{status_color.get(r['status'], '#333')};font-weight:600'>"
        f"{r['status'].upper()}</td>"
        f"<td>{r['duration_s']}s</td></tr>"
        for r in sorted(results, key=lambda x: x["nodeid"])
    )
    html = f"""<!doctype html>
<html><head><meta charset="utf-8">
<title>NutriScan AI — Selenium Execution Report</title>
<style>
body {{ font-family: -apple-system, Segoe UI, Roboto, sans-serif; margin: 2rem; color: #1a1a1a; }}
h1 {{ margin-bottom: 0.25rem; }}
.meta {{ color: #666; margin-bottom: 1.5rem; }}
.cards {{ display: flex; gap: 1rem; margin-bottom: 2rem; }}
.card {{ padding: 1rem 1.5rem; border-radius: 10px; background: #f4f4f4; min-width: 120px; }}
.card b {{ display:block; font-size: 1.6rem; }}
table {{ border-collapse: collapse; width: 100%; }}
th, td {{ text-align: left; padding: 0.5rem 0.75rem; border-bottom: 1px solid #eee; font-size: 0.9rem; }}
th {{ background: #fafafa; position: sticky; top: 0; }}
</style></head>
<body>
<h1>NutriScan AI — Selenium Execution Report</h1>
<p class="meta">Target: {config.BASE_URL} &middot; Generated: {datetime.now(timezone.utc).isoformat()}</p>
<div class="cards">
  <div class="card">Total<b>{summary['passed'] + summary['failed'] + summary['skipped']}</b></div>
  <div class="card" style="background:#eaf7ee">Passed<b style="color:#1e8e4a">{summary['passed']}</b></div>
  <div class="card" style="background:#fdeceb">Failed<b style="color:#d64545">{summary['failed']}</b></div>
  <div class="card" style="background:#fdf3d9">Skipped<b style="color:#b58900">{summary['skipped']}</b></div>
  <div class="card">Pass rate<b>{summary['pass_rate']}%</b></div>
  <div class="card">Duration<b>{summary['total_duration_s']}s</b></div>
</div>
<table>
<thead><tr><th>Test</th><th>Status</th><th>Duration</th></tr></thead>
<tbody>
{rows}
</tbody>
</table>
</body></html>
"""
    with open(out_path, "w", encoding="utf-8") as fh:
        fh.write(html)


def write_dashboard_html(summary, out_path):
    by_module = summary["by_module"]
    bars = ""
    for mod, counts in sorted(by_module.items()):
        total = sum(counts.values()) or 1
        passed_pct = counts.get("passed", 0) / total * 100
        failed_pct = counts.get("failed", 0) / total * 100
        skipped_pct = counts.get("skipped", 0) / total * 100
        short_mod = mod.split("/")[-1].replace(".py", "")
        bars += f"""
<div style="margin-bottom:0.75rem">
  <div style="font-size:0.85rem;margin-bottom:0.2rem">{short_mod} ({total})</div>
  <div style="display:flex;height:14px;border-radius:7px;overflow:hidden;background:#eee">
    <div style="width:{passed_pct}%;background:#1e8e4a"></div>
    <div style="width:{failed_pct}%;background:#d64545"></div>
    <div style="width:{skipped_pct}%;background:#b58900"></div>
  </div>
</div>"""
    html = f"""<!doctype html>
<html><head><meta charset="utf-8"><title>NutriScan AI — Test Dashboard</title>
<style>
body {{ font-family: -apple-system, Segoe UI, Roboto, sans-serif; margin: 2rem; }}
h1 {{ margin-bottom: 0.25rem; }}
.big {{ font-size: 3rem; font-weight: 700; }}
.gate-pass {{ color: #1e8e4a; }}
.gate-fail {{ color: #d64545; }}
</style></head>
<body>
<h1>NutriScan AI — Test Dashboard</h1>
<p class="big {'gate-pass' if summary['pass_rate'] >= config.PASS_RATE_THRESHOLD else 'gate-fail'}">
  {summary['pass_rate']}%
</p>
<p>Gate: {config.PASS_RATE_THRESHOLD}% required to pass CI &middot;
   {summary['passed']} passed / {summary['failed']} failed / {summary['skipped']} skipped &middot;
   {summary['total_duration_s']}s total</p>
<h2>By module</h2>
{bars}
</body></html>
"""
    with open(out_path, "w", encoding="utf-8") as fh:
        fh.write(html)


# ── helpers ──────────────────────────────────────────────────────────────────

def _make_header(ws, columns, fill_color=_HDR):
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center")
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
    ws.freeze_panes = "A2"


def _row_fields(r: dict) -> dict:
    ids = _parse_nodeid(r["nodeid"])
    func_source = _SRC_CACHE.get(ids["filepath"], ids["funcname"], ids["classname"])

    primary_marker = (r.get("markers") or "").split(",")[0].strip()
    category, severity = _MARKER_INFO.get(primary_marker, ("General", "Medium"))

    title = "Verify " + _humanize(ids["funcname"]).rstrip(".")
    if ids["param_id"]:
        title += f" (variant: {ids['param_id']})"
    target = _target_route(func_source, ids["param_id"])
    description = f"{title} ({category} flow). Target: {target}."

    return {
        "category": category,
        "title": title,
        "description": description,
        "target": target,
        "severity": severity,
    }


def write_xlsx(results, summary, out_path, run_at=None):
    """Produce Automation_Test_Report.xlsx's 'Selenium E2E' sheet: one row
    per passing test, columns matching the reference per-test-case QA
    report format."""
    if Workbook is None:
        print("openpyxl not installed — skipping .xlsx report", file=sys.stderr)
        return

    passed = sorted(
        (r for r in results if r["status"] == "passed"),
        key=lambda x: x["nodeid"],
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Selenium E2E"
    _make_header(ws, [
        "Test ID", "Category", "Test Case Title", "Test Description",
        "Target URL/Selector", "Severity/Priority", "Status",
        "Response/Execution Time (ms)", "Details",
    ])
    for seq, r in enumerate(passed, start=1):
        f = _row_fields(r)
        duration_ms = round(r["duration_s"] * 1000, 1)
        details = f"Completed in {duration_ms}ms against {f['target']}."
        ws.append([
            f"SEL-{seq:03d}", f["category"], f["title"], f["description"],
            f["target"], f["severity"], "PASSED", duration_ms, details,
        ])
        row_idx = ws.max_row
        ws.cell(row=row_idx, column=7).fill = PatternFill("solid", fgColor=_STATUS_CELL_COLORS["PASSED"])
        for col in range(1, 10):
            ws.cell(row=row_idx, column=col).alignment = Alignment(vertical="center", wrap_text=True)
    for col_letter, width in zip("ABCDEFGHI", [10, 22, 40, 60, 44, 14, 10, 18, 55]):
        ws.column_dimensions[col_letter].width = width

    wb.save(out_path)
    print(f"Wrote {out_path}")


def main():
    os.makedirs(config.REPORTS_DIR, exist_ok=True)
    results, workers, files = load_all_results()

    if not results:
        print("WARNING: no result files found under "
              f"{config.RESULTS_DIR} — did pytest run?", file=sys.stderr)

    run_at = datetime.now(timezone.utc).isoformat()
    summary = summarize(results)
    print(f"Merged {len(files)} worker file(s): {workers}")
    print(f"Total={summary['total']} Passed={summary['passed']} "
          f"Failed={summary['failed']} Skipped={summary['skipped']} "
          f"PassRate={summary['pass_rate']}% Duration={summary['total_duration_s']}s")

    write_execution_results_json(
        results, summary,
        os.path.join(config.REPORTS_DIR, "execution-results.json"),
        run_at=run_at,
    )
    write_summary_md(summary, os.path.join(config.REPORTS_DIR, "summary.md"))
    write_execution_report_html(
        results, summary, os.path.join(config.REPORTS_DIR, "execution-report.html")
    )
    write_dashboard_html(summary, os.path.join(config.REPORTS_DIR, "dashboard.html"))
    write_xlsx(
        results, summary,
        os.path.join(config.REPORTS_DIR, "Automation_Test_Report.xlsx"),
        run_at=run_at,
    )

    print("Reports written to", config.REPORTS_DIR)


if __name__ == "__main__":
    main()
